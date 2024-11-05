import PySimpleGUI as sg
import Sub_works_references
import Sub_works_writing
import DOCX_writing
import final_schedule
import pandas as pd
import datetime
import os
import requests
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def apply_format_to_sheet(sheet):
    """Apply formatting to the sheet based on cell content."""
    for row in sheet.iter_rows():
        for cell in row:
            content = cell.value
            if isinstance(content, str) and '$#$' in content:
                parts = content.split('$#$')
                num_value = float(parts[1]) if len(parts) > 2 else float(parts[-1])
                is_italic = parts[2] == 'False' if len(parts) > 2 else False
                cell.fill = get_fill_color(num_value)
                if is_italic:
                    cell.font = Font(underline='single')

def get_fill_color(value):
    """Get the color for cell based on value."""
    if 0.6 <= value < 0.75:
        intensity = int((value - 0.6) * 255 / 0.15)
        return PatternFill(start_color=f"FF{255-intensity:02X}{255-intensity:02X}", end_color=f"FF{255-intensity:02X}{255-intensity:02X}", fill_type="solid")
    elif 0.75 <= value < 0.9:
        intensity = int((value - 0.75) * 255 / 0.15)
        return PatternFill(start_color=f"FFFF{intensity:02X}00", end_color=f"FFFF{intensity:02X}00", fill_type="solid")
    elif 0.9 <= value <= 1:
        intensity = int((value - 0.9) * 255 / 0.1)
        return PatternFill(start_color=f"FF{intensity:02X}FF00", end_color=f"FF{intensity:02X}FF00", fill_type="solid")
    
    # Default case to avoid returning None
    return PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")


def create_folder_if_not_exists(folder_name):
    """Create folder if it doesn't exist."""
    if not os.path.exists(folder_name):
        os.mkdir(folder_name)
        print(f"Created folder: {folder_name}")


def download_model_if_not_exists(model_url, model_path):
    """Download model if it does not exist locally."""
    if not os.path.exists(model_path):
        response = requests.get(model_url)
        if response.status_code == 200:
            with open(model_path, "wb") as model_file:
                model_file.write(response.content)
            print(f"Model downloaded and saved to {model_path}")
        else:
            print(f"Failed to download the model. Status code: {response.status_code}")


def merge_and_align_cells(workbook, processed_data, path):
    """Merge and align cells in the workbook and save it."""
    sheet = workbook.active
    cols = list(processed_data.columns)
    for col_num, col in enumerate(cols, start=1):
        start_row = None
        for row_num in range(2, sheet.max_row + 2):
            if row_num <= sheet.max_row and sheet.cell(row=row_num, column=col_num).value == sheet.cell(row=row_num-1, column=col_num).value:
                if start_row is None:
                    start_row = row_num - 1
            else:
                if start_row:
                    sheet.merge_cells(start_row=start_row, start_column=col_num, end_row=row_num-1, end_column=col_num)
                    start_row = None
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    workbook.save(path)


def main():
    work_folders = [
        'main_work_comparision',
        'sub_work_comparision',
        'Main_work_schedule_report',
        'sub_work_schedule_report',
    ]

    for folder in work_folders:
        create_folder_if_not_exists(os.path.join(os.getcwd(), folder))

    model_url = "https://huggingface.co/lmstudio-community/Meta-Llama-3.1-8B-Instruct-GGUF/raw/main/Meta-Llama-3.1-8B-Instruct-Q8_0.gguf"
    model_dir = "models"
    model_filename = "Meta-Llama-3.1-8B-Instruct-Q8_0.gguf"
    download_model_if_not_exists(model_url, os.path.join(model_dir, model_filename))

    layout = [
        [sg.Checkbox("Use AI", key="-USE_AI-")],
        [sg.Button("Process primary_bid", key="-LOA1-"), sg.Button("Process sub_bid", key="-MIN_LOA-")],
        [sg.Text("Select the main work comparision sheet:"), sg.Input(key="-EXCEL1-"), sg.FileBrowse()],
        [sg.Text("Select the sub work comparision sheet:"), sg.Input(key="-EXCEL2-"), sg.FileBrowse()],
        [sg.Button("Process final main schedule", key="-FINAL_main_work-")],
        [sg.Button("Process final sub schedule", key="-FINAL_sub_work-")],
    ]
    
    window = sg.Window("Tender Rate Comparator", layout)
    current_directory= os.getcwd()

    while True:
        event, values = window.read()
        use_AI = int(values.get("-USE_AI-", 0))
        if(use_AI ==1):
            print('\nUsing AI model\n')
            use_AI =1
        else:
            print('\n Not using AI model\n')
            use_AI =0
        if event == sg.WINDOW_CLOSED:
            break
        elif event == "-LOA1-":
            try:
                processed_data1 = final_schedule.main(use_AI) 
                toime = datetime.datetime.now().strftime('%H_%M_%S of %d-%m')
                file_name = f"temporary_file_{toime}.xlsx"
                folder_path = os.path.join(current_directory,work_folders[0])
                path1 = os.path.join(folder_path, file_name)
                #print('Path1 taken')
                processed_data1.to_excel(path1, engine = 'openpyxl')
                file_name = f"main_works_comparision_file_{toime}.xlsx"
                patha = os.path.join(folder_path, file_name)
                workbook = load_workbook(path1)
                for sheet_name in workbook.sheetnames:
                    apply_format_to_sheet(workbook[sheet_name])
                # Save the workbook after formatting
                workbook.save(patha)
                sg.popup('The Excel sheet named primary_work_bid_comparision is generated in ',work_folders[0])
            except:
                sg.popup('Please select valid HTML/PDF files')
        elif event == "-MIN_LOA-":
            try:
                processed_data2 = Sub_works_references.main(use_AI)
                toime = datetime.datetime.now().strftime('%H_%M_%S of %d-%m')
                file_name = f"temporary_file_{toime}.xlsx"
                folder_path = os.path.join(current_directory,work_folders[1])
                path2 = os.path.join(folder_path, file_name)
                excel_writer = pd.ExcelWriter(path2, engine="openpyxl")
                # Write each DataFrame to a separate sheet in the Excel file
                for idx, df in enumerate(processed_data2, start=1):
                    sheet_name = f"Subwork_{idx}"  # Change the sheet name as needed
                    df.to_excel(excel_writer, sheet_name=sheet_name, index=False)
                excel_writer.book.save(path2)
                workbook = load_workbook(path2)
                for sheet_name in workbook.sheetnames:
                    apply_format_to_sheet(workbook[sheet_name])
                file_name = f"sub_works_comparision_file_{toime}.xlsx"
                pathb = os.path.join(folder_path, file_name)
                workbook.save(pathb)
                sg.popup('The excel sheet is generated with name sub_work_bid_comparision on the desktop')
            except:
                sg.popup('Please select valid sub work HTML/PDF files')
        elif event == "-FINAL_main_work-":
            try:
                filepath = values["-EXCEL1-"]
                if filepath:
                    processed_data3 = DOCX_writing.main(filepath)
                    toime = datetime.datetime.now().strftime('%H_%M_%S of %d-%m')
                    file_name = f"Final_main_Schedule_report_{toime}.xlsx"
                    folder_path = os.path.join(current_directory,work_folders[2])
                    path3 = os.path.join(folder_path, file_name)
                    processed_data3.to_excel(path3, engine = 'openpyxl')
                    workbook = load_workbook(path3)
                    sheet = workbook.active
                    cols = list(processed_data3.columns)
                    for col_num, col in enumerate(cols, start=1):
                        start_row = None
                        for row_num in range(2, sheet.max_row + 2):
                            if row_num <= sheet.max_row and sheet.cell(row=row_num, column=col_num).value == sheet.cell(row=row_num-1, column=col_num).value:
                                if start_row is None:
                                    start_row = row_num - 1
                            else:
                                if start_row:
                                    sheet.merge_cells(start_row=start_row, start_column=col_num, end_row=row_num-1, end_column=col_num)
                                    start_row = None
                    for row in sheet.iter_rows():
                        for cell in row:
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    workbook.save(path3)
                    sg.popup('The excel sheet is generated with name Final_Schedule_report on the desktop')
            except:
                sg.popup('Please select correct main work Excelsheet files')
        elif event == "-FINAL_sub_work-":
            try:
                filepath = values["-EXCEL2-"]
                if filepath:
                    toime = datetime.datetime.now().strftime('%H_%M_%S of %d-%m')
                    file_name = f"Final_sub_Schedule_report_{toime}.xlsx"
                    folder_path = os.path.join(current_directory,work_folders[3])
                    path4 = os.path.join(folder_path, file_name)
                    processed_data4 = Sub_works_writing.main(filepath)
                    processed_data4.to_excel(path4, engine = 'openpyxl')
                    workbook = load_workbook(path4)
                    sheet = workbook.active
                    cols = list(processed_data4.columns)
                    for col_num, col in enumerate(cols, start=1):
                        start_row = None
                        for row_num in range(2, sheet.max_row + 2):  # Adjusting the loop to consider one more iteration
                            if row_num <= sheet.max_row and sheet.cell(row=row_num, column=col_num).value == sheet.cell(row=row_num-1, column=col_num).value:
                                if start_row is None:
                                    start_row = row_num - 1
                            else:
                                if start_row:
                                    sheet.merge_cells(start_row=start_row, start_column=col_num, end_row=row_num-1, end_column=col_num)
                                    start_row = None
                    for row in sheet.iter_rows():
                        for cell in row:
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    workbook.save(path4)
                    sg.popup('The excel sheet is generated with name Final_Sub_schedule_report on the desktop')
            except:
                sg.popup('Please select correct subwork Excelsheet files')
    window.close()
if __name__ == "__main__":
    main()
        
